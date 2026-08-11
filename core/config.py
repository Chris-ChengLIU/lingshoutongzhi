"""
配置管理模块
管理群组列表、消息模板的读写和持久化
"""

import json
import os
import csv
from pathlib import Path
from typing import List, Dict, Optional


# 获取数据目录路径
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data"
LOGS_DIR = BASE_DIR / "logs"

# 确保目录存在
DATA_DIR.mkdir(exist_ok=True)
LOGS_DIR.mkdir(exist_ok=True)

GROUPS_FILE = DATA_DIR / "groups.json"
TEMPLATES_FILE = DATA_DIR / "templates.json"


class GroupConfig:
    """群组配置管理"""

    def __init__(self, file_path: Path = GROUPS_FILE):
        self.file_path = file_path
        self.groups: List[Dict] = []
        self.load()

    def load(self):
        """从文件加载群组列表"""
        if self.file_path.exists():
            try:
                with open(self.file_path, "r", encoding="utf-8") as f:
                    self.groups = json.load(f)
            except (json.JSONDecodeError, IOError):
                self.groups = []
        else:
            self.groups = []

    def save(self):
        """保存群组列表到文件"""
        with open(self.file_path, "w", encoding="utf-8") as f:
            json.dump(self.groups, f, ensure_ascii=False, indent=2)

    def add_group(self, name: str, webhook: str, category: str = "默认") -> Dict:
        """添加群组"""
        group = {
            "id": len(self.groups) + 1,
            "name": name,
            "webhook": webhook,
            "category": category,
            "enabled": True
        }
        # 重新分配ID
        if self.groups:
            group["id"] = max(g["id"] for g in self.groups) + 1
        self.groups.append(group)
        self.save()
        return group

    def update_group(self, group_id: int, name: str = None, webhook: str = None,
                     category: str = None, enabled: bool = None) -> bool:
        """更新群组信息"""
        for group in self.groups:
            if group["id"] == group_id:
                if name is not None:
                    group["name"] = name
                if webhook is not None:
                    group["webhook"] = webhook
                if category is not None:
                    group["category"] = category
                if enabled is not None:
                    group["enabled"] = enabled
                self.save()
                return True
        return False

    def delete_group(self, group_id: int) -> bool:
        """删除群组"""
        for i, group in enumerate(self.groups):
            if group["id"] == group_id:
                self.groups.pop(i)
                self.save()
                return True
        return False

    def get_group(self, group_id: int) -> Optional[Dict]:
        """获取单个群组"""
        for group in self.groups:
            if group["id"] == group_id:
                return group
        return None

    def get_enabled_groups(self) -> List[Dict]:
        """获取所有启用的群组"""
        return [g for g in self.groups if g.get("enabled", True)]

    def get_categories(self) -> List[str]:
        """获取所有分类"""
        categories = set(g.get("category", "默认") for g in self.groups)
        return sorted(categories)

    def import_from_csv(self, file_path: str) -> tuple:
        """
        从CSV文件批量导入群组
        CSV格式: 群名称, webhook地址, 分类(可选)
        返回: (成功数, 失败数, 错误信息列表)
        """
        success_count = 0
        fail_count = 0
        errors = []

        try:
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                header = next(reader, None)

                for line_num, row in enumerate(reader, start=2):
                    try:
                        if len(row) < 2:
                            errors.append(f"第{line_num}行: 数据不完整")
                            fail_count += 1
                            continue

                        name = row[0].strip()
                        webhook = row[1].strip()
                        category = row[2].strip() if len(row) > 2 and row[2].strip() else "默认"

                        if not name or not webhook:
                            errors.append(f"第{line_num}行: 群名称或webhook为空")
                            fail_count += 1
                            continue

                        self.add_group(name, webhook, category)
                        success_count += 1
                    except Exception as e:
                        errors.append(f"第{line_num}行: {str(e)}")
                        fail_count += 1

        except Exception as e:
            errors.append(f"读取文件失败: {str(e)}")

        return success_count, fail_count, errors

    def import_from_excel(self, file_path: str) -> tuple:
        """
        从Excel文件批量导入群组
        Excel格式: 第一列群名称, 第二列webhook地址, 第三列分类(可选)
        返回: (成功数, 失败数, 错误信息列表)
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return 0, 0, ["需要安装 openpyxl 库才能读取 Excel 文件"]

        success_count = 0
        fail_count = 0
        errors = []

        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            for line_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or len(row) < 2:
                        errors.append(f"第{line_num}行: 数据不完整")
                        fail_count += 1
                        continue

                    name = str(row[0]).strip() if row[0] else ""
                    webhook = str(row[1]).strip() if row[1] else ""
                    category = str(row[2]).strip() if len(row) > 2 and row[2] else "默认"

                    if not name or not webhook:
                        errors.append(f"第{line_num}行: 群名称或webhook为空")
                        fail_count += 1
                        continue

                    self.add_group(name, webhook, category)
                    success_count += 1
                except Exception as e:
                    errors.append(f"第{line_num}行: {str(e)}")
                    fail_count += 1

            wb.close()

        except Exception as e:
            errors.append(f"读取文件失败: {str(e)}")

        return success_count, fail_count, errors

    def export_to_csv(self, file_path: str) -> bool:
        """导出群组列表为CSV"""
        try:
            with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["群名称", "Webhook地址", "分类", "启用状态"])
                for group in self.groups:
                    writer.writerow([
                        group["name"],
                        group["webhook"],
                        group.get("category", "默认"),
                        "是" if group.get("enabled", True) else "否"
                    ])
            return True
        except Exception:
            return False

    def export_to_excel(self, file_path: str) -> bool:
        """导出群组列表为Excel"""
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "群组列表"

            # 写入表头
            ws.append(["群名称", "Webhook地址", "分类", "启用状态"])

            # 写入数据
            for group in self.groups:
                ws.append([
                    group["name"],
                    group["webhook"],
                    group.get("category", "默认"),
                    "是" if group.get("enabled", True) else "否"
                ])

            # 调整列宽
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 10

            wb.save(file_path)
            return True
        except Exception:
            return False


class TemplateConfig:
    """消息模板管理"""

    def __init__(self, file_path: Path = TEMPLATES_FILE):
        self.file_path = file_path
        self.templates: List[Dict] = []
        self.load()

    def load(self):
        """从文件加载模板列表"""
        if self.file_path.exists():
            try:
                with open(self.file_path, "r", encoding="utf-8") as f:
                    self.templates = json.load(f)
            except (json.JSONDecodeError, IOError):
                self.templates = []
        else:
            self.templates = []
            # 添加默认模板
            self._add_default_templates()

    def _add_default_templates(self):
        """添加默认模板"""
        default_templates = [
            {
                "id": 1,
                "name": "会议通知",
                "content": """<font color="info">【会议通知】</font>

> **会议主题：** {topic}
> **会议时间：** {time}
> **会议地点：** {location}
> **参会人员：** {attendees}

{notes}"""
            },
            {
                "id": 2,
                "name": "培训通知",
                "content": """<font color="warning">【培训通知】</font>

> **培训主题：** {topic}
> **培训时间：** {time}
> **培训地点：** {location}
> **参训人员：** {attendees}

{notes}"""
            },
            {
                "id": 3,
                "name": "紧急通知",
                "content": """<font color="warning">【紧急通知】</font>

{content}"""
            }
        ]
        self.templates = default_templates
        self.save()

    def save(self):
        """保存模板列表到文件"""
        with open(self.file_path, "w", encoding="utf-8") as f:
            json.dump(self.templates, f, ensure_ascii=False, indent=2)

    def add_template(self, name: str, content: str) -> Dict:
        """添加模板"""
        template = {
            "id": max([t["id"] for t in self.templates], default=0) + 1,
            "name": name,
            "content": content
        }
        self.templates.append(template)
        self.save()
        return template

    def update_template(self, template_id: int, name: str = None, content: str = None) -> bool:
        """更新模板"""
        for template in self.templates:
            if template["id"] == template_id:
                if name is not None:
                    template["name"] = name
                if content is not None:
                    template["content"] = content
                self.save()
                return True
        return False

    def delete_template(self, template_id: int) -> bool:
        """删除模板"""
        for i, template in enumerate(self.templates):
            if template["id"] == template_id:
                self.templates.pop(i)
                self.save()
                return True
        return False

    def get_template(self, template_id: int) -> Optional[Dict]:
        """获取单个模板"""
        for template in self.templates:
            if template["id"] == template_id:
                return template
        return None

    def get_template_names(self) -> List[str]:
        """获取所有模板名称"""
        return [t["name"] for t in self.templates]


# ── 追加内容开始 ────────────────────────────────────────────────────────────

PERSONS_FILE = DATA_DIR / "persons.json"


class PersonConfig:
    """
    联系人配置管理

    数据结构（persons.json）：
    [
      {
        "id": 1,
        "name": "张三",
        "unit": "信用卡中心",
        "selector": null,   // null = 选第1条; "2" = 选第2条
        "enabled": true
      },
      ...
    ]
    """

    def __init__(self, file_path: Path = PERSONS_FILE):
        self.file_path = file_path
        self.persons: List[Dict] = []
        self.load()

    def load(self):
        """从文件加载联系人列表"""
        if self.file_path.exists():
            try:
                with open(self.file_path, "r", encoding="utf-8") as f:
                    self.persons = json.load(f)
            except (json.JSONDecodeError, IOError):
                self.persons = []
        else:
            self.persons = []

    def save(self):
        """保存联系人列表到文件"""
        with open(self.file_path, "w", encoding="utf-8") as f:
            json.dump(self.persons, f, ensure_ascii=False, indent=2)

    def _next_id(self) -> int:
        return max((p["id"] for p in self.persons), default=0) + 1

    def add_person(
        self,
        name: str,
        unit: str = "",
        tag="",
        # selector: Optional[str] = None,
    ) -> Dict:
        """添加联系人"""
        person = {
            "id":       self._next_id(),
            "name":     name,
            "unit":     unit,
            "tag":      tag,
            # "selector": selector,   # None 或数字字符串 "2"、"3"…
            "enabled":  True,
        }
        self.persons.append(person)
        self.save()
        return person

    def update_person(
        self,
        person_id: int,
        name: str = None,
        unit: str = None,
        tag:  str = None,
        # selector: Optional[str] = None,
        enabled: bool = None,
    ) -> bool:
        """更新联系人信息"""
        for p in self.persons:
            if p["id"] == person_id:
                if name     is not None: p["name"]     = name
                if unit     is not None: p["unit"]      = unit
                if tag is not None: p["tag"] = tag
                # if selector is not None: p["selector"]  = selector
                if enabled  is not None: p["enabled"]   = enabled
                self.save()
                return True
        return False

    def delete_person(self, person_id: int) -> bool:
        """删除联系人"""
        for i, p in enumerate(self.persons):
            if p["id"] == person_id:
                self.persons.pop(i)
                self.save()
                return True
        return False

    def get_person(self, person_id: int) -> Optional[Dict]:
        """获取单个联系人"""
        return next((p for p in self.persons if p["id"] == person_id), None)

    def get_enabled_persons(self) -> List[Dict]:
        """获取所有启用的联系人"""
        return [p for p in self.persons if p.get("enabled", True)]

    def get_by_unit(self, unit: str) -> List[Dict]:
        """
        按所在单位筛选启用的联系人。
        unit 为空字符串或 "全部" 时返回所有启用联系人。
        """
        if not unit or unit == "全部":
            return self.get_enabled_persons()
        return [p for p in self.persons if p.get("enabled", True) and p.get("unit") == unit]

    def get_units(self) -> List[str]:
        """获取所有不重复的所在单位，按字母排序"""
        units = sorted(set(p.get("unit", "") for p in self.persons if p.get("unit")))
        return units
    
    def get_tags(self) -> List[str]:
        """拆分所有 person 的 tag 字段，去重排序"""
        all_tags = set()
        for p in self.persons:
            for t in (p.get("tag") or "").split(","):
                t = t.strip()
                if t:
                    all_tags.add(t)
        return sorted(all_tags)

    def get_by_tag(self, tag: str) -> List[Dict]:
        if not tag or tag == "全部":
            return self.get_enabled_persons()
        return [
            p for p in self.persons
            if p.get("enabled", True)
            and tag in [t.strip() for t in (p.get("tag") or "").split(",")]
        ]
    
    def import_from_csv(self, file_path: str) -> tuple:
        """
        从 CSV 批量导入联系人。

        CSV 表头（第一行）需包含：姓名、所在单位、自定义标签、选择序号
        - 姓名：必填
        - 所在单位、自定义标签、选择序号：均可选，可以不写这一列，也可以调换顺序

        Returns:
            (success_count, fail_count, errors)
        """
        success_count = 0
        fail_count    = 0
        errors: List[str] = []

        try:
            with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)

                for line_num, row in enumerate(reader, start=2):
                    try:
                        # 过滤空行（所有字段都是空的情况）
                        if not any((v or "").strip() for v in row.values()):
                            continue

                        name = (row.get("姓名") or "").strip()
                        if not name:
                            errors.append(f"第{line_num}行: 姓名为空")
                            fail_count += 1
                            continue

                        unit = (row.get("所在单位") or "").strip()
                        tag  = (row.get("自定义标签") or "").strip()


                        self.add_person(name, unit, tag)
                        success_count += 1

                    except Exception as e:
                        errors.append(f"第{line_num}行: {str(e)}")
                        fail_count += 1

        except Exception as e:
            errors.append(f"读取文件失败: {str(e)}")

        return success_count, fail_count, errors


    def import_from_excel(self, file_path: str) -> tuple:
        """
        从 Excel 批量导入联系人。

        第一行为表头，需包含：姓名、所在单位、自定义标签、选择序号
        （除"姓名"外均可选，顺序任意）

        Returns:
            (success_count, fail_count, errors)
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return 0, 0, ["需要安装 openpyxl 库才能读取 Excel 文件"]

        success_count = 0
        fail_count    = 0
        errors: List[str] = []

        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            # 读表头，建立 列名 -> 列索引 的映射
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                wb.close()
                return 0, 0, ["Excel 文件为空或没有表头"]

            col_index = {}
            for idx, col_name in enumerate(header_row):
                if col_name:
                    col_index[str(col_name).strip()] = idx

            def cell(row, col_name):
                idx = col_index.get(col_name)
                if idx is None or idx >= len(row) or row[idx] is None:
                    return ""
                return str(row[idx]).strip()

            for line_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or not any((v is not None and str(v).strip()) for v in row):
                        continue

                    name = cell(row, "姓名")
                    if not name:
                        errors.append(f"第{line_num}行: 姓名为空")
                        fail_count += 1
                        continue

                    unit = cell(row, "所在单位")
                    tag  = cell(row, "自定义标签")


                    self.add_person(name, unit, tag)
                    success_count += 1

                except Exception as e:
                    errors.append(f"第{line_num}行: {str(e)}")
                    fail_count += 1

            wb.close()

        except Exception as e:
            errors.append(f"读取文件失败: {str(e)}")

        return success_count, fail_count, errors


    def export_to_csv(self, file_path: str) -> bool:
        """导出联系人列表为 CSV"""
        try:
            with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["姓名", "所在单位", "自定义标签",  "启用状态"])
                for p in self.persons:
                    writer.writerow([
                        p["name"],
                        p.get("unit", ""),
                        p.get("tag", ""),
                        "是" if p.get("enabled", True) else "否",
                    ])
            return True
        except Exception:
            return False

    def clear_all(self):
        """清空所有联系人"""
        self.persons.clear()
        self.save()